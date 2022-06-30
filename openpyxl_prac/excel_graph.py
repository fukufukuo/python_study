# エクセルのグラブ描画を行う

# モジュールインポート
import openpyxl
import pandas as pd
from datetime import datetime # datetimeモジュールからdatetime関数をimport

# 変数定義
read_path = R"C:\Python_study\openpyxl_prac\data\shokki_stock_data.xlsx"
output_name = "C:\Python_study\openpyxl_prac\output\output.xlsx"

# データ読み取り
df = pd.read_excel(read_path, sheet_name=0, usecols='B,D') # date列とhigh_value列のみ読み込み

# グラフ描画エクセル作成
wb = openpyxl.Workbook(output_name)
ws = wb.create_sheet(title='うんちman')

# エクセル保存
wb.save(output_name) # すでに存在する状態で再度実行すると、ファイルが上書きされる

# エクセル読み込み
wb = openpyxl.load_workbook(output_name)
ws = wb.worksheets[0]

# index記入
ws['A1'] = 'date'
ws['B1'] = '最高値'

# dfからエクセルにデータ転記
for i in range(len(df)): # range(stop)はstopを超えない数を0から順番に返す len(df)でdfの行数を取得
  ws['A' + str(i+2)] = datetime.strftime(df.iloc[i,0], '%Y-%m-%d') # iloc[a,b]でa行b列の要素を返す strptime(文字列、整形書式)で、文字列をdatetimeオブジェクトに変換
  ws['B' + str(i+2)] = df.iloc[i,1]

# グラフ描画に用いる関数をimport
from openpyxl.chart import Reference
from openpyxl.chart.axis import DateAxis

# グラフオブジェクト生成
graph_obj = openpyxl.chart.LineChart() # 折れ線グラフ 散布図ならScatterChart()
graph_obj.title = 'ショッキ株価高値推移'




# エクセル保存
wb.save(output_name)
