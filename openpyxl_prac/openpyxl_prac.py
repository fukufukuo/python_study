# openpyxlの練習

import openpyxl

# 変数定義
read_path = r'C:\Python_study\openpyxl_prac\data\shokki_stock_data.xlsx'
wb = openpyxl.load_workbook(read_path)
ws = wb.worksheets[0]
cell_value = []

# iter_rows()の練習
# セルの内包表記を用いて、2行目から5行目まで、一行ごとにリストを作成し、それらをまとめたリストを作る
for row in ws.iter_rows(min_row=2, max_row=5):
  cell_value.append([cell.value for cell in row]) #内包することで、ループした要素が並ぶリストを作れる 確かにこれはシンプルでいい

print(cell_value)

